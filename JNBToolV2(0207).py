
from PyQt5.QtCore import QFile, QIODevice, Qt, QTextStream
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import (QDialog, QFileDialog, QGridLayout, QHBoxLayout, QMessageBox,
		QLabel, QLineEdit, QPushButton, QTextEdit, QVBoxLayout, QComboBox, QRadioButton, QCheckBox,
		QWidget)
import os
import shutil
import openpyxl
from openpyxl.styles import Font

from openpyxl import styles
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side, PatternFill, colors
import random
import datetime
import webbrowser
import ctypes
import requests
myappid = 'mycompany.myproduct.subproduct.version' # arbitrary string
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)


dateSubmit = str(datetime.date.today())
print(dateSubmit)

listDesAV = {' Choose a designer... ':'','Trung': '중A', 'Kieu': '끼우', 
			'Thuy': '튀B', 'Chinh': '찐', 'Hai': '하이B', 'Dong': '돔', 'Quan': '타잉', 'Thinh': '꾸언'}
ListDesV = list(listDesAV.keys())
ListDesV.sort()
#autofic Function
### CHUOI HAM SU LI FILE MDB

def saveAsMdbFile(oldname,newname ):
	old = open(oldname, 'rb') 
	content = old.read()
	new = open(newname, 'wb')
	new.write(content)
	new.close()

## CHUỖI HÀM XỬ LÍ FILE EXCEL

def autoCorrectEx(filename, toa, kv, dateSub = dateSubmit):
	# Load mot workBook
	# wbNamwe = filename
	wb = openpyxl.load_workbook(filename)
	sheet = wb.active

	# Lay gioi han cua bang tinh
	max_row = sheet.max_row
	max_column = sheet.max_column

	if max_column != 14:
		raise TypeError
	else:
		pass

	# Nhap ten toa va khu vu
	tenToa = str(toa).upper()
	khuVuc = (str(kv)).upper()

	# dat ten cho bang thong ke (title)
	sheet.merge_cells('A2:N2')
	titFont = Font(bold = True, size = 20)
	sheet['A2'].font = titFont
	sheet['A2'].alignment = Alignment(horizontal = 'center')

	title = "물량집계표(%s 동%s_계단)" %(tenToa.upper(), khuVuc.upper())
	sheet['A2'] = title

	#Sua ten toa
	sheet['L5'].value = tenToa
	sheet['L6'].value = khuVuc

	# Sua project name cell
	boldFont = Font(bold = True, size = 11)
	sheet['A4'].font = boldFont
	sheet['A4'].alignment = Alignment(horizontal = 'left')

	# Sua date cell
	sheet['N4'].font = boldFont
	sheet['N4'].alignment = Alignment(horizontal = 'right')
	sheet['N4'].value = dateSub

	# Sua Head row
	thin_border = Border(left=Side(style='thin'), 
	                     right=Side(style='thin'), 
	                     top=Side(style='thin'), 
	                     bottom=Side(style='thin'))

	for col in 'ABCDEFGHIJKLMN':
		for row in range(5, 7):
			sheet[str(col) + str(row)].font = boldFont
			sheet[str(col) + str(row)].alignment = Alignment(horizontal = 'center', vertical = 'center')
			sheet[str(col) + str(row)].border = thin_border

	# Sua total row
	for col in 'ABCDEFGHIJKLMN':
		sheet[str(col) + str(max_row)].font = boldFont

	# merge cells 
	sheet.merge_cells('K5:K6')

	print("01.Dinh dang file: OK")

	### PART2: EDIT CONTENT:

	# SUA STT TRONG EXCEL:
	for row in range(7, max_row + 1):
		row = str(row)
		sheet['F' + row].value = sheet['A' + row].value
	print("02.Sua cot STT : OK")

	# Sua ten tam:
	platesWrong = ['SDH', 'SP0', 'SSP', 'D(']

	for row in range(7, max_row + 1):
		row = str(row)
		cellVal = str(sheet['B' + row].value)
		if cellVal[0:3] in platesWrong or cellVal[0:2] in platesWrong:
			if cellVal[0:2] == 'SP':
				sheet['B' + row].value = cellVal[:2] + cellVal[-3:]
			elif cellVal[0:2] == 'D(':
				sheet['B' + row].value = cellVal[:1] + cellVal[-3:]
			else:
				sheet['B' + row].value = cellVal[:3] + cellVal[-3:]
		else:
			pass

	# Sua dien tich SA, SCP

	AREA1Wrong = ['SA-', 'SCP']

	for row in range(7, max_row + 1):
		row = str(row)
		cellValName = str(sheet['B' + row].value)
		cellValVol = (sheet['I' + row].value)
		cellValArea = (sheet['G' + row].value)
		if cellValName[0:3] in AREA1Wrong and (cellValVol < 3.5 or cellValArea < 0.1):
			if cellValName[0:2] == 'SA':
				sheet['G' + row].value = round(int(cellValName[3:6]) * int(cellValName[-4:])/1000000, 3)
				sheet['I' + row].value = sheet['G' + row].value * 35
			else:
				pass
				sheet['G' + row].value = round(int(cellValName[4:7]) * int(cellValName[-4:])/1000000, 3)
				sheet['I' + row].value = sheet['G' + row].value * 15

				# sheet['G' + row].value = cellValName[4:7] + cellValName[-4:]
				# print("SCPs fixed")

			# print(cellValName)
		else:
			pass

	# Sua dien tich va khoi luong SDH, SP, SSP

	AREA2Wrong = ['SDH', 'SP-', 'SSP']

	for row in range(7, max_row + 1):
		row = str(row)
		cellValName = str(sheet['B' + row].value)
		cellValVol = (sheet['I' + row].value)
		cellValArea = (sheet['G' + row].value)

		if cellValName[0:3] in AREA2Wrong :
			if cellValName[0:2] == 'SP' and (cellValVol < 3.5 or cellValArea < 0.1):
				sheet['G' + row].value = round(random.uniform(0.564, 0.695), 3)
				sheet['I' + row].value = sheet['G' + row].value * 17

			elif cellValName[0:3] == 'SDH' and cellValName[-1] == '1' and (cellValVol < 2 or cellValArea < 0.1):
				sheet['G' + row].value = round(random.uniform(0.191, 0.235), 3)
				sheet['I' + row].value = sheet['G' + row].value * 35

			elif cellValName[0:3] == 'SDH' and cellValName[-1] == '2' and (cellValVol < 1.5 or cellValArea < 0.05):
				sheet['G' + row].value = round(random.uniform(0.065, 0.101), 3)
				sheet['I' + row].value = sheet['G' + row].value * 54

			elif cellValName[0:3] == 'SSP' and (cellValVol < 3.5 or cellValArea < 0.1):
				sheet['G' + row].value = round(random.uniform(0.891, 1.095), 3)
				sheet['I' + row].value = sheet['G' + row].value * 20

			else:
				pass
		else:
			pass

	### PART 3: DINH DANG IN AN.

	# Dinh dang cot
	wN = sheet.column_dimensions['N']
	wN.width = 55

	# Dinh dang in
	sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

	print("04.Dinh dang in A4: OK")
	# Luu sang file khac

	saveName = "autoCorrected Output File.xlsx"
	wb.save(saveName)

	# Mo file sau khi sua
	# os.startfile(saveName)
	return saveName

#autofic Function

def saveFileEx(fileIn, info, fileOut = "None"):
	# Load mot workBook
	wbI = openpyxl.load_workbook(fileIn)
	sheet = wbI.active

	# Nhan ban excel(sua toa, title):

	title = "물량집계표(%s 동%s_계단)" %(info[0].upper(), info[1].upper())
	sheet['A2'] = title

	# Sua ten toa:
	sheet['L5'].value = info[0]
	sheet['L6'].value = info[1]

	# Sua ngay trinh:
	sheet['N4'].value = info[5]

	# Sua Head row
	thin_border = Border(left=Side(style='thin'), 
	                     right=Side(style='thin'), 
	                     top=Side(style='thin'), 
	                     bottom=Side(style='thin'))

	noneFill = PatternFill(start_color='FFFF0000',
	                   end_color='FFFF0000',
	                   fill_type='none')	

	for col in 'ABCDEFGHIJKLMN':
		for row in range(5, 7):
			sheet[str(col) + str(row)].border = thin_border

	for row in sheet.iter_rows():
		for cell in row:
			cell.fill = noneFill

	# Save file
	wbI.save(fileOut)
	# Mo file sau khi sua
	os.startfile(fileOut)


# Compare 2 files:

def compareEx(fileIn, fileOut):
	wbI = openpyxl.load_workbook(fileIn)
	wbO = openpyxl.load_workbook(fileOut)
	wsI = wbI.active
	wsO = wbO.active

	max_row = wsI.max_row
	max_col = wsI.max_column

		# Warn!!
	greenFill = PatternFill(start_color='91189431',
	                   end_color='91189431',
	                   fill_type='solid')

	for row in wsI.iter_rows():
		for cell in row:
			col = str(cell.column)
			row = str(cell.row)
			if wsI[str(col + row)].value == wsO[str(col + row)].value:
				pass
			else:
				wsO[str(col + row)].fill = greenFill
	wbO.save("outCompare.xlsx")
	os.startfile("outCompare.xlsx")


# cHUỖI HÀM XỬ LÍ FILE MOL

# Update file MOL, BOM


def updateMOL(mol, kl ='KL.xlsx',modeBj = True):

	# Mo file MOL len va copy noi dung cua file ra 1 list input

	molInputFile = open(mol, 'r')
	listInputMol = []
	for line in molInputFile:
		if line.count('|') == 7:
			listInputMol.append(line)
	molInputFile.close()

	# Load file excel mau
	klwb = openpyxl.load_workbook(kl, data_only = True)
	IPws = klwb['INPUT']
	CODEws = klwb['PASTE CODE']
	listPaste = []

	updateMolFile = open(mol,'w+')

	# Load file excel Mau, lay paste code, 

	listPlate = []
	for row in range(2, CODEws.max_row):
		row = str(row)
		if (str(CODEws['B' + row].value)).count('|') == 7:
			updateMolFile.write(str(CODEws['B' + row].value)+ "\n")

			listRow = str(CODEws['B' + row].value).split('|')
			if (listRow[1] + listRow[2]) in listPlate:
				pass
			else:
				listPlate.append(listRow[1] + listRow[2])
		else:
			pass

	# Luon them listplate BJ350

	listPlate.append('BJ350')

	# Da tao duoc list plate already existed


	# load more from original .MOL:

	for line in listInputMol:
		match = line.split('|')
		if (match[1] + match[2]) in listPlate:
			pass
		else:
			updateMolFile.writelines(line) # Ghi file mol
	if modeBj == True:
		updateMolFile.writelines("\n000000BJ035003500000000000000000000+00000|BJ350|||AF-BJ001|350||1")
		print("Included BJ350")
	else:
		print('Not Included BJ350')
		pass
		
	updateMolFile.close()
	# os.startfile(mol)

def updateBom(bomFile, kl ='KL.xlsx', modeBj = True):

	# Mo file Bom len va copy noi dung cua file ra 1 list input

	BomInputFile = open(bomFile, 'r')
	listInputBom = []

	for line in BomInputFile:
		if line.count(",") == 35:
			listInputBom.append(line.split(','))
	BomInputFile.close()

	# Tới đây chúng ta được một danh sách các tấm nguyên bản từ file BOM

	# Load file excel mau
	klwb = openpyxl.load_workbook(kl, data_only = True)
	IPws = klwb['INPUT']
	CODEws = klwb['PASTE CODE']

	# Tao 1 dict match:

	dictMatch = {}

	for row in range(2, CODEws.max_row + 3):
		row = str(row)
		if (str(CODEws['F' + row].value)).count('|') == 1:
			subList = str(CODEws['F' + row].value).split('|')
			dictMatch[subList[0]] = str(subList[1])
	# print(dictMatch)
	# Tới đây chúng ta được 1 dic lưu {tên tấm: Diện tích}


	# Đếm số tấm DP = 1/2 Số tấm BJ

	amountDp = 0
	for i in listInputBom:
		if (i[1]+i[2]+i[3]) in dictMatch.keys():
			i[6] = dictMatch[str(i[1]+i[2]+i[3])]
		elif i[1] == 'DP-':
			amountDp +=1
		elif i[1]+i[2] == 'BJ350':
			del listInputBom[listInputBom.index(i)]
		else:
			pass	
	# print(listInputBom[0])
	# print("So tam Dp co la: %d" % amountDp)
	try:
		block = listInputBom[0][0]
	except:
		print("BOM file is empty")
	amountBj350 = str(2*amountDp)
	rowBj350 = [block, 'BJ', '350', '', '', amountBj350, '', 'ALFORM', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'ALFORM', 'BJ', 'M2', 'H', 'BJ01\n']
	# print(rowBj350)

	if modeBj:
		if rowBj350 not in listInputBom:
			listInputBom.append(rowBj350)
			
			print("Included BJ350")
		else:
			print("The BJ350 already exists")
	else:
		if rowBj350 in listInputBom:
			print("chỉ số BJ350 = " + list.index(rowBj350))
		print("Not Included BJ")
		pass		
	# for i in listInputBom:
	# 	if 
	# 	print(i)


	# Ghi File:

	outBomFile = open(bomFile, 'w+')
	try:
		for i in listInputBom:
			line = ','.join(i)
			outBomFile.write(line)
	except:
		print("Loi! Khong them duoc BJ")
	finally :
		outBomFile.close()



# Tao UI:

class EditExcelFile(QWidget):

	filename = "None"
	report = "None"


	def __init__(self, parent = None):
		super(EditExcelFile, self).__init__(parent)

		#Icon Window:
		self.setWindowIcon(QIcon("Icon.ico"))
		self.setWindowTitle("JNB Tool V1")

		# Set mode:
		self.modeBom = True
		self.modeMdb = False
		self.modeEx = False

		self.isBjchecked = True
		# UI:
		
		# Other button:
		# Movie Button
		self.warnLine = QTextEdit()
		self.warnLine.setFixedSize(450, 75)
		self.warnLine.setText("""<html><b><span style="color: red">WARNING:</span><BR>
- Tool chỉ dùng cho các trường hợp thông thường.<BR>
- Các mẫu bảng biểu, tên tấm theo mẫu hiện tại (2018-02-03).<BR>
- ..... !!!</b></html> """)

		self.clipButton = QPushButton('&Link F...')
		self.clipButton.setMinimumHeight(30)
		self.clipButton.setToolTip('18+.')

		self.docButton = QPushButton('&Link V.')
		self.docButton.setMinimumHeight(30)
		self.docButton.setToolTip('Black&White Beautiful Girl  <3.')


		self.helpLayout = QHBoxLayout()
		self.helpLayout.addWidget(self.warnLine)
		self.helpLayout.addSpacing(20)
		self.helpLayout.addWidget(self.clipButton)
		self.helpLayout.addWidget(self.docButton)
		self.helpLayout.addSpacing(50)
		# self.helpLayout.addSpacing(50)
		# self.helpLayout.addStretch()


		# self.helpMainLayout = QHBoxLayout()
		# self.helpMainLayout.addLayout(self.helpLayout)
		# self.helpMainLayout.addSpacing(200)

		# Choose mode:

		# Open data setting (command BOM cad)
		self.openDataSet = QPushButton('''Open Data Setting''')
		self.openDataSet.setMinimumHeight(80)
		self.openDataSet.setToolTip('Run Data Setting.')

		# check box include BJ350

		self.incBjCheckBox = QCheckBox("BJ350")
		self.incBjCheckBox.setEnabled(True)
		self.incBjCheckBox.setChecked(True)
		self.incBjCheckBox.setEnabled(True)
		self.amountBj = QLineEdit()
		self.amountBj.setEnabled(False)

		self.incBjLayout = QHBoxLayout()
		self.incBjLayout.addWidget(self.incBjCheckBox)
		self.incBjLayout.addWidget(self.amountBj)

		self.incBjCheckBox.stateChanged.connect(self.checkBj)

		# Choose mode layout

		self.modeLayout = QVBoxLayout()
		self.modeLayout.addWidget(self.openDataSet)
		
		self.modeLayout.addSpacing(45)
		
		self.modeRB = QRadioButton('Edit BOM, MOL File')
		self.modeRB.setChecked(True)
		self.modeRB.mode = "Edit BOM, MOL File" 
		self.modeRB.toggled.connect(self.setMode)
		self.modeLayout.addWidget(self.modeRB)
		self.modeLayout.addLayout(self.incBjLayout)

		self.modeLayout.addSpacing(58)

		
		self.modeRB = QRadioButton('Edit MDB File')
		self.modeRB.mode = "Edit MDB File" 
		self.modeRB.toggled.connect(self.setMode)
		self.modeLayout.addWidget(self.modeRB)
		self.modeLayout.addSpacing(27)

		
		self.modeRB = QRadioButton('Edit Excel File')
		self.modeRB.mode = "Edit Excel File" 
		self.modeRB.toggled.connect(self.setMode)
		self.modeLayout.addWidget(self.modeRB)

		self.modeLayout.addSpacing(63)

		self.modeLayout.addStretch()




		# Edit Excel File:

		self.editExlable = QLabel("Edit Excel File:")
		self.editExlable.setEnabled(self.modeEx)
		self.pathExLable = QLabel('Path Excel File:')
		self.pathExLine = QLineEdit()
		self.pathExLine.setReadOnly(True)
		self.pathExLine.setPlaceholderText("Nhấn 'Load Excel File...' để lựa chọn file Excel cần chỉnh sửa.")
		self.pathExLine.setEnabled(self.modeEx)

		self.loadExButton = QPushButton('&Load Excel File ...')
		self.loadExButton.setToolTip('Load input file .xlsx from harddisk.')
		self.loadExButton.setEnabled(self.modeEx)

		self.correctExButton = QPushButton('AutoCorrect')
		self.correctExButton.setToolTip('AutoCorrect input file xlsx.')
		self.correctExButton.setEnabled(self.modeEx)

		self.saveExButton = QPushButton('Save')
		self.saveExButton.setToolTip('Save the file xlsx.')
		self.saveExButton.setEnabled(self.modeEx)


		# Edit Mdb File:

		self.editMdblable = QLabel("Edit MDB File:")
		self.editMdblable.setEnabled(self.modeEx)
		self.pathMdbLable = QLabel('Path MDB File:')
		self.pathMdbLine = QLineEdit()
		self.pathMdbLine.setReadOnly(True)
		self.pathMdbLine.setPlaceholderText("Nhấn 'Load MDB File...' để lựa chọn file MDB cần chỉnh sửa.")
		self.pathMdbLine.setEnabled(self.modeEx)

		self.loadMdbButton = QPushButton('&Load MDB File ...')
		self.loadMdbButton.setToolTip('Load input file .MDBfrom harddisk.')
		self.loadMdbButton.setEnabled(self.modeEx)

		self.correctMdbButton = QPushButton('Manually Edit')
		self.correctMdbButton.setToolTip('Edit input file mdb with Access.')
		self.correctMdbButton.setEnabled(self.modeEx)

		self.saveMdbButton = QPushButton('Save')
		self.saveMdbButton.setToolTip('Save the file mdb.')
		self.saveMdbButton.setEnabled(self.modeMdb)



		# Info:

		infoLable = QLabel('General Infor:')
		toaLable = QLabel('Block:')
		self.toaLine = QLineEdit()
		self.toaLine.setReadOnly(False)
		self.toaLine.setPlaceholderText("Tên tòa...")

		kvLable = QLabel('Zone: ') # kv = Khu vuc
		self.kvLine = QLineEdit()
		self.kvLine.setReadOnly(False)
		self.kvLine.setPlaceholderText("Tên khu vực...")

		
		dateLable= QLabel('Date (yyyy-mm-dd):')
		self.dateLine = QLineEdit()
		self.dateLine.setPlaceholderText("Ngày trình...")
		self.dateLine.setText(dateSubmit)


		Des2Lable= QLabel('Designer:')
		self.Des2Combo = QComboBox()
		self.Des2Combo.addItems(ListDesV)



		DesLable= QLabel('Designer:')
		self.DesLine = QLineEdit()
		self.DesLine.setPlaceholderText("Tên Designer...")

		#Edit .MOL:

		self.pathKlLable = QLabel('KL file:')
		self.pathKlLine = QLineEdit()
		self.pathKlLine.setReadOnly(True)
		self.pathKlLine.setPlaceholderText("Nhấn 'Load File KL...' để chọn file.")

		self.loadKlButton = QPushButton('&Load KL File...')
		self.loadKlButton.setToolTip('Load input file .kl from harddisk.')

		self.editKlButton = QPushButton('&Edit KL File...')
		self.editKlButton.setToolTip('Edit file .kl')

		self.pathMolLable = QLabel('MOL file:')
		self.pathMolLine = QLineEdit()
		self.pathMolLine.setReadOnly(True)
		self.pathMolLine.setPlaceholderText("Nhấn 'Load File MOL...' để chọn file.")

		self.loadMolButton = QPushButton('&Load MOL File...')
		self.loadMolButton.setToolTip('Load input file .MOL from harddisk.')

		self.updateMolButton = QPushButton('&Update MOL...')
		self.updateMolButton.setToolTip('Update file MOL.')


		self.pathBomLable = QLabel('BOM file:')
		self.pathBomLine = QLineEdit()
		self.pathBomLine.setReadOnly(True)
		self.pathBomLine.setPlaceholderText("Nhấn 'Load File Bom...' để chọn file.")

		self.loadBomButton = QPushButton('&Load Bom File...')
		self.loadBomButton.setToolTip('Load input file .BOM from harddisk.')

		self.updateBomButton = QPushButton('&Update BOM...')
		self.updateBomButton.setToolTip('Update file BOM.')



		statusLable = QLabel("Notice:")
		self.statusBox = QTextEdit()
		report = "Các thông báo trạng thái sẽ xuất hiện ở đây!"
		self.statusBox.setText(report)
		self.statusBox.setReadOnly(True)

		self.checkButton = QPushButton('Check')
		self.checkButton.setToolTip('Highlight the Error cells.')
		self.checkButton.hide()

		self.quitButton = QPushButton('Quit')
		self.quitButton.setToolTip('Quit only.')

		self.quitReButton = QPushButton('Remove temp files.')
		self.quitReButton.setToolTip('Remove temp files and Quit.')

		# layout

		buttonLayout1 = QVBoxLayout()
		buttonLayout1.addStretch()
		buttonLayout1.addWidget(self.quitReButton)
		buttonLayout1.addWidget(self.quitButton)


		#info Layout

		infoLayout = QGridLayout()
		infoLayout.addWidget(toaLable, 0, 0)
		infoLayout.addWidget(kvLable, 1, 0)
		infoLayout.addWidget(dateLable, 2, 0)
		infoLayout.addWidget(DesLable, 3, 0)
		infoLayout.addWidget(self.toaLine, 0, 1)
		infoLayout.addWidget(self.kvLine, 1, 1)
		infoLayout.addWidget(self.dateLine, 2, 1)
		infoLayout.addWidget(self.DesLine, 3, 1)
		infoLayout.addWidget(self.Des2Combo, 3, 2)


		#Edit Excel Layout
		editExLayout = QHBoxLayout()
		editExLayout.addWidget(self.pathExLable)
		editExLayout.addWidget(self.pathExLine)
		editExLayout.addWidget(self.loadExButton)
		editExLayout.addWidget(self.correctExButton)
		editExLayout.addWidget(self.saveExButton)


		#Edit Excel Layout
		editMdbLayout = QHBoxLayout()
		editMdbLayout.addWidget(self.pathMdbLable)
		editMdbLayout.addWidget(self.pathMdbLine)
		editMdbLayout.addWidget(self.loadMdbButton)
		editMdbLayout.addWidget(self.correctMdbButton)
		editMdbLayout.addWidget(self.saveMdbButton)

		# Edit MOL layout:
		
		#Edit kl Layout
		editKlLayout = QHBoxLayout()
		editKlLayout.addWidget(self.pathKlLable)
		editKlLayout.addWidget(self.pathKlLine)
		editKlLayout.addWidget(self.loadKlButton)
		editKlLayout.addWidget(self.editKlButton)

		#update mol Layout
		updateMolLayout = QHBoxLayout()
		updateMolLayout.addWidget(self.pathMolLable)
		updateMolLayout.addWidget(self.pathMolLine)
		updateMolLayout.addWidget(self.loadMolButton)
		updateMolLayout.addWidget(self.updateMolButton)

		#update bom Layout

		updateBomLayout = QHBoxLayout()
		updateBomLayout.addWidget(self.pathBomLable)
		updateBomLayout.addWidget(self.pathBomLine)
		updateBomLayout.addWidget(self.loadBomButton)
		updateBomLayout.addWidget(self.updateBomButton)

		# Edit MOL layout

		editMolLayout = QVBoxLayout()
		# editMolLayout.addWidget(editMolLable)
		# editMolLayout.addWidget(self.molModeRB)
		editMolLayout.addLayout(editKlLayout)
		editMolLayout.addLayout(updateMolLayout)
		editMolLayout.addLayout(updateBomLayout)


		# Report Layout
		NoticeLayout = QVBoxLayout()
		NoticeLayout.addWidget(statusLable)
		NoticeLayout.addWidget(self.statusBox)


		# Main 1 layout:
		mainVLayout = QVBoxLayout()
		mainVLayout.addWidget(infoLable)
		mainVLayout.addLayout(infoLayout)
		mainVLayout.addSpacing(20)
		# mainVLayout.addLayout(self.modeLayout)
		mainVLayout.addSpacing(10)

		mainVLayout.addLayout(editMolLayout)

		mainVLayout.addSpacing(20)
		mainVLayout.addLayout(editMdbLayout)

		mainVLayout.addSpacing(20)
		mainVLayout.addLayout(editExLayout)
		mainVLayout.addSpacing(20)
		mainVLayout.addLayout(NoticeLayout)
		mainVLayout.addStretch()


		mainLayout = QGridLayout()
		mainLayout.addLayout(mainVLayout, 0, 0)
		mainLayout.addLayout(self.modeLayout, 0, 1)
		mainLayout.addLayout(self.helpLayout, 1, 0)
		mainLayout.addLayout(buttonLayout1, 1, 1)


		self.setLayout(mainLayout)

		# Connect
		# info connect
		self.Des2Combo.currentIndexChanged.connect(self.chooseDes)


		# info connect
		self.openDataSet.clicked.connect(self.runDataSet)


		# Radio Button:
		# self.molModeRB.ton

		#Edit excel connect
		self.loadExButton.clicked.connect(self.loadExClick)
		self.saveExButton.clicked.connect(self.saveExOut)
		self.correctExButton.clicked.connect(self.correctExFilesClick)

		#Edit MDB connect
		self.loadMdbButton.clicked.connect(self.loadMdbClick)
		self.saveMdbButton.clicked.connect(self.saveMdb)
		self.correctMdbButton.clicked.connect(self.runOutMdbFile)

		#Edit kl connect
		self.loadKlButton.clicked.connect(self.loadKlClick)
		self.editKlButton.clicked.connect(self.editKlClick)

		#Edit mol connect
		self.loadMolButton.clicked.connect(self.loadMolClick)
		self.updateMolButton.clicked.connect(self.updateMolClick)

		#Edit mol connect
		self.loadBomButton.clicked.connect(self.loadBomClick)
		self.updateBomButton.clicked.connect(self.updateBomClick)




		#Help connect
		self.clipButton.clicked.connect(self.clipClicked)
		self.docButton.clicked.connect(self.docClicked)


		#Quit connect
		self.quitButton.clicked.connect(self.quitClicked)
		self.quitReButton.clicked.connect(self.quitReClicked)

		# Set Mode Default

		self.pathKlLable.setEnabled(self.modeBom)
		self.pathKlLine.setEnabled(self.modeBom)
		self.loadKlButton.setEnabled(self.modeBom)
		self.editKlButton.setEnabled(self.modeBom)

		self.pathMolLable.setEnabled(self.modeBom)
		self.pathMolLine.setEnabled(self.modeBom)
		self.loadMolButton.setEnabled(self.modeBom)
		self.updateMolButton.setEnabled(self.modeBom)

		self.pathBomLable.setEnabled(self.modeBom)
		self.pathBomLine.setEnabled(self.modeBom)
		self.loadBomButton.setEnabled(self.modeBom)
		self.updateBomButton.setEnabled(self.modeBom)
		self.incBjCheckBox.setEnabled(self.modeBom)



		report = ''' <html><b>CÁC THÔNG BÁO TRẠNG THÁI IN RA TẠI ĐÂY!!!</b><BR>
					\nĐang chọn 'Edit BOM, MOL File' Mode, có các chức năng:<BR>
					\n* Chỉnh sửa các file MOL, file BOM.<BR>
					\n* Cập nhật diện tích, khối lượng các tấm lỗi từ file KL*.xlsx.<BR>
					\n ===> <b><span style="color: red">Chú ý: Các thao tác update sẽ ảnh hưởng trực tiếp lên các file, hãy cẩn thận!.</span></b></html>
					'''
		self.statusBox.setText(report)
		print("Ban chon mode BOM")

		
	# HÀM SET MODE

	def setMode(self):

		self.modeRB = self.sender()


		if self.modeRB.isChecked():
			if self.modeRB.mode == 'Edit BOM, MOL File':
				self.modeEx = False
				self.editExlable.setEnabled(self.modeEx)
				self.pathExLine.setEnabled(self.modeEx)
				self.loadExButton.setEnabled(self.modeEx)
				self.saveExButton.setEnabled(self.modeEx)
				self.correctExButton.setEnabled(self.modeEx)

				self.modeBom = True

				self.pathKlLable.setEnabled(self.modeBom)
				self.pathKlLine.setEnabled(self.modeBom)
				self.loadKlButton.setEnabled(self.modeBom)
				self.editKlButton.setEnabled(self.modeBom)

				self.pathMolLable.setEnabled(self.modeBom)
				self.pathMolLine.setEnabled(self.modeBom)
				self.loadMolButton.setEnabled(self.modeBom)
				self.updateMolButton.setEnabled(self.modeBom)

				self.pathBomLable.setEnabled(self.modeBom)
				self.pathBomLine.setEnabled(self.modeBom)
				self.loadBomButton.setEnabled(self.modeBom)
				self.updateBomButton.setEnabled(self.modeBom)
				self.incBjCheckBox.setEnabled(self.modeBom)

				self.modeMdb = False
				self.editMdblable.setEnabled(self.modeMdb)
				self.pathMdbLine.setEnabled(self.modeMdb)
				self.loadMdbButton.setEnabled(self.modeMdb)
				self.saveMdbButton.setEnabled(self.modeMdb)
				self.correctMdbButton.setEnabled(self.modeMdb)	

				report = ''' <html>Đang chọn 'Edit BOM, MOL File' Mode, có các chức năng:<BR>
							\n* Chỉnh sửa các file MOL, file BOM.<BR>
							\n* Cập nhật diện tích, khối lượng các tấm lỗi từ file KL*.xlsx.<BR>
							\n ===> <b><span style="color: red">Chú ý: Các thao tác update sẽ
							ảnh hưởng trực tiếp lên các file, hãy cẩn thận!.</span></b></html>
							'''
				self.statusBox.setText(report)
				print("Ban chon mode BOM")



			elif self.modeRB.mode == "Edit Excel File" :

				self.modeEx = True
				self.editExlable.setEnabled(self.modeEx)
				self.pathExLine.setEnabled(self.modeEx)
				self.loadExButton.setEnabled(self.modeEx)
				self.saveExButton.setEnabled(self.modeEx)
				self.correctExButton.setEnabled(self.modeEx)

				self.modeMdb = False
				self.editMdblable.setEnabled(self.modeMdb)
				self.pathMdbLine.setEnabled(self.modeMdb)
				self.loadMdbButton.setEnabled(self.modeMdb)
				self.saveMdbButton.setEnabled(self.modeMdb)
				self.correctMdbButton.setEnabled(self.modeMdb)	

				self.modeBom = False
				self.pathKlLable.setEnabled(self.modeBom)
				self.pathKlLine.setEnabled(self.modeBom)
				self.loadKlButton.setEnabled(self.modeBom)
				self.editKlButton.setEnabled(self.modeBom)


				self.pathMolLable.setEnabled(self.modeBom)
				self.pathMolLine.setEnabled(self.modeBom)
				self.loadMolButton.setEnabled(self.modeBom)
				self.updateMolButton.setEnabled(self.modeBom)

				self.pathBomLable.setEnabled(self.modeBom)
				self.pathBomLine.setEnabled(self.modeBom)
				self.loadBomButton.setEnabled(self.modeBom)
				self.updateBomButton.setEnabled(self.modeBom)

				self.incBjCheckBox.setEnabled(self.modeBom)

				report = ''' Đang chọn 'Edit Excel File' Mode, có các chức năng:
							\n* Định dạng title, header row, date cell, total row, trang in A4.
							\n* Sửa tên các tấm lỗi.
							\n* Sửa cột số thứ tự.
							\n* Các ô được chỉnh sửa thay đổi sẽ được bôi xanh.
							\n* Tự động lưu theo tên file excel theo mẫu.
							\n ===> Chú ý: Kiểm tra, sửa lại nếu cần thiết.
							'''							
				self.statusBox.setText(report)


				print("Ban chon mode EXCEL")
			else:
				self.modeEx = False
				self.editExlable.setEnabled(self.modeEx)
				self.pathExLine.setEnabled(self.modeEx)
				self.loadExButton.setEnabled(self.modeEx)
				self.saveExButton.setEnabled(self.modeEx)
				self.correctExButton.setEnabled(self.modeEx)


				self.modeMdb = True
				self.editMdblable.setEnabled(self.modeMdb)
				self.pathMdbLine.setEnabled(self.modeMdb)
				self.loadMdbButton.setEnabled(self.modeMdb)
				self.saveMdbButton.setEnabled(self.modeMdb)
				self.correctMdbButton.setEnabled(self.modeMdb)	


				self.modeBom = False
				self.pathKlLable.setEnabled(self.modeBom)
				self.pathKlLine.setEnabled(self.modeBom)
				self.loadKlButton.setEnabled(self.modeBom)
				self.editKlButton.setEnabled(self.modeBom)

				self.pathMolLable.setEnabled(self.modeBom)
				self.pathMolLine.setEnabled(self.modeBom)
				self.loadMolButton.setEnabled(self.modeBom)
				self.updateMolButton.setEnabled(self.modeBom)

				self.pathBomLable.setEnabled(self.modeBom)
				self.pathBomLine.setEnabled(self.modeBom)
				self.loadBomButton.setEnabled(self.modeBom)
				self.updateBomButton.setEnabled(self.modeBom)

				self.incBjCheckBox.setEnabled(self.modeBom)

				report = ''' Đang chọn 'Edit MDB File' Mode, có các chức năng:
							\n* Sửa tay file MDB bằng Access.
							\n* Lưu tên theo mẫu.
							\n ===> Chú ý: Phải cài Microsoft Office Access vào mới sửa được :).
							'''
				self.statusBox.setText(report)

				print("ban chon mode MOL")
				


	# HÀM XỦ LÍ FILE MOL, BOM, KL

	def loadKlClick(self):
		fileName, _ = QFileDialog.getOpenFileName(self, "Open the file",
		'', "Excel file (*.xlsx);;All Files (*)")

		if fileName == "":
			report = " Bạn chưa chọn file."
		else:
			report = '''
						\nFile: '%s' được load thành công!
						\n- Hãy điền đầy đủ các thông tin bên dưới.
						\n- Nhấn nút "Edit KL File..." để thực hiện việc chỉnh sửa bổ sung bằng tay.
						''' % fileName.upper()

		self.pathKlLine.setText(fileName)
		self.statusBox.setText(report)
		print(fileName)
		self.editKlClick()

	def editKlClick(self):
		print("EdKlButton Clicked")
		if self.pathKlLine.text() == "":
			report = " Bạn chưa chọn file muốn chỉnh sửa, hãy nhấn 'Load KL File....' để chọn."
		else:
			os.startfile(self.pathKlLine.text())
			report = "Mở file %s thành công." % self.pathKlLine.text()
		self.statusBox.setText(report)


	def loadMolClick(self):
		fileName, _ = QFileDialog.getOpenFileName(self, "Open the file",
		'', "Excel file (*.mol);;All Files (*)")

		if fileName == "":
			report = " Bạn chưa chọn file muốn chỉnh sửa, hãy nhấn 'Load MOL File....' để chọn."
		else:
			report = '''
						\nFile: '%s' được load thành công!
						\n- Hãy điền đầy đủ các thông tin bên dưới.
						\n- Nhấn nút "Update MOL.." để thực hiện việc update tự động.
						''' % fileName.upper()
		self.statusBox.setText(report)

		self.pathMolLine.setText(fileName)
		print(fileName)

	def updateMolClick(self):

		kl = self.pathKlLine.text()
		mol = self.pathMolLine.text()
		print(kl)
		print(mol)

		report = "Bạn đã chọn chức năng sửa file MOL"
		try:
			updateMOL(mol, kl, self.isBjchecked)
			report = "Update file MOL thành công."		
		except PermissionError:
			report = "ERROR! Có thể file '...MOL' đang được mở, hãy đóng và thử lại."
		except FileNotFoundError:
			report = '''
			\nERROR! Hãy chọn file cần chỉnh sửa bằng "Load File MOL...".'
			\nERROR! Hãy chọn file cần chỉnh sửa bằng "Load KL File...".'
			'''

		except NameError:
			report = 'ERROR! Hãy nhập tên toàn và khu vực.'
		
		# self.statusBox.setText(report)
		self.statusBox.setText(report)
		print("updateMolButton Clicked")


	def loadBomClick(self):
		fileName, _ = QFileDialog.getOpenFileName(self, "Open the file",
		'', "Excel file (*.Bom);;All Files (*)")

		if fileName == "":
			report = " Bạn chưa chọn file muốn chỉnh sửa, hãy nhấn 'Load BOM File....' để chọn."
		else:
			report = '''
						\nFile: '%s' được load thành công!
						\n- Hãy điền đầy đủ các thông tin bên dưới.
						\n- Nhấn nút "Update BOM.." để thực hiện việc update tự động.
						''' % fileName.upper()
		self.statusBox.setText(report)

		self.pathBomLine.setText(fileName)
		print(fileName)

	def updateBomClick(self):

		kl = self.pathKlLine.text()
		Bom = self.pathBomLine.text()
		print(kl)
		print(Bom)

		report = "Bạn đã chọn chức năng sửa file Bom"
		try:
			updateBom(Bom, kl, self.isBjchecked)
			report = "Update file Bom thành công."		
		except PermissionError:
			report = "ERROR! Có thể file '...Bom' đang được mở, hãy đóng và thử lại."
		except FileNotFoundError:
			report = '''
			\nERROR! Hãy chọn file cần chỉnh sửa bằng "Load File Bom...".'
			\nERROR! Hãy chọn file cần chỉnh sửa bằng "Load KL File...".'
			'''

		# except NameError:
		# 	report = 'ERROR! Xảy ra lỗi không xác định, file BOM chưa được update.'

		except:
			report = '''ERROR! file BOM chưa được update. Các lỗi có thể gặp phải:
			\n* File BOM vừa load vào là file rỗng.
			\n* File BOM vừa load vào chứa kí tự hoặc định dạng đặc biệt.
			'''
		
		# self.statusBox.setText(report)
		self.statusBox.setText(report)
		print("updateBomButton Clicked")


	def checkBj(self):
		report = ''
		if self.incBjCheckBox.isChecked():
			self.isBjchecked = True
			report = "Bạn đã CHỌN thêm BJ350 tự động. Hãy nhấn 'updateMOL...' và 'updateBOM...' để cập nhật lại file."
			print("Bj checked")
		else:
			self.isBjchecked = False
			report = "Bạn đã BỎ CHỌN thêm BJ350 tự động.\n* Hãy nhấn 'updateMOL...' và 'updateBOM...' để cập nhật lại file.\n* Các tấm BJ350 sẽ KHÔNG có trong MOL, BOM."
			print("Bj not checked")

		self.statusBox.setText(report)


	# HÀM XỬ EDIT MDB

	def loadMdbClick(self):
		fileName, _ = QFileDialog.getOpenFileName(self, "Open the file",
		'', "Access file (*.mdb);;All Files (*)")

		if fileName == "":
			report = " Bạn chưa chọn file muốn chỉnh sửa, hãy nhấn 'Load Mdb File....' để chọn."
		else:
			report = '''
						\nFile: '%s' được load thành công!
						\n- Hãy điền đầy đủ các thông tin bên trên.
						\n- Nhấn nút "Manually Edit" để thực hiện việc chỉnh sửa thủ công trong Access.
						''' % fileName.upper()
		self.statusBox.setText(report)

		self.pathMdbLine.setText(fileName)
		print(fileName)

	def saveMdbOut(self):
		try:
			self.info = (self.toaLine.text().upper(), self.kvLine.text().upper(),self.DesLine.text() ,self.dateLine.text()[-5:-3], self.dateLine.text()[-2:], self.dateLine.text())
			print(self.info[5])
			fileNameSave, _ = QFileDialog.getSaveFileName(self,"saveFlle","%s동%s_계단_%s(%s%s).mdb" % self.info[0:5] ,filter ="File Mdb (*.mdb *.)")
			fileAC = self.pathMdbLine.text()
			saveFileMdb(fileAC,self.info, fileNameSave)
			report = 'OK! File được lưu thành công.'
		except PermissionError:
			report = "ERROR! Có thể file %s đang mở, hãy đóng và thử lại lần nữa." % fileNameSave
		# except TypeError:
		# 	report = 'Bạn đã hủy việc lưu file. File lưu không thành công.'

		except FileNotFoundError:
			report = 'ERROR! Hãy thực hiện việc sửa file bằng cách nhấn "AutoCorrect" trước khi lưu file.'

		self.statusBox.setText(report)
		# print(fileNameSave)

	def saveMdb(self):
		try:
			self.info = (self.toaLine.text().upper(), self.kvLine.text().upper(),self.DesLine.text() ,self.dateLine.text()[-5:-3], self.dateLine.text()[-2:], self.dateLine.text())
			print(self.info[5])
			stringName = "%s동%s_계단_%s(%s%s).mdb" % self.info[0:5]
			fileNameSave, _ = QFileDialog.getSaveFileName(self,"saveFileMdb","%s동%s_계단_%s(%s%s).mdb" % self.info[0:5] ,filter ="File Mdb (*.mdb *.)")
			saveAsMdbFile(self.pathMdbLine.text(), fileNameSave)
			self.pathMdbLine.setText(str(fileNameSave))
			report = 'OK! File "%s" được lưu thành công.' %fileNameSave
		except PermissionError:
			report = "ERROR! Có thể file %s đang mở, hãy đóng và thử lại lần nữa." % fileNameSave
		except TypeError:
			report = 'Bạn đã hủy việc lưu file. File lưu không thành công.'

		except FileNotFoundError:
			report = 'ERROR! Hãy thực hiện việc sửa file bằng cách nhấn "Manually Edit" trước khi lưu file.'

		self.statusBox.setText(report)
		# print(fileNameSave)


	def runOutPro(self):
		os.startfile('HD-FORM2015License(Admin)-CRACK.exe')

	def runOutMdbFile(self):
		try:

			if self.pathMdbLine.text() == '':
				raise FileNotFoundError()
			os.startfile(self.pathMdbLine.text())
			
			report = 'Mở file thành công'
		except PermissionError:
			report = "ERROR! Có thể file %s đang mở, hãy đóng và thử lại lần nữa." % fileNameSave
		# except TypeError:
		# 	report = 'Bạn đã hủy việc lưu file. File lưu không thành công.'

		except FileNotFoundError:
			report = 'ERROR! Bạn chưa chọn file mdb để sửa, chọn file bằng cách nhấn "Load MDB File...".'

		self.statusBox.setText(report)

	def autoCorrectFileMdb(self):
		filename = self.pathMdbLine.text()
		toa = self.toaLine.text()
		kv = self.kvLine.text()
		dateSub = self.dateLine.text()
		
		if toa =='':
			print("Check")

		try:
			fO = autoCorrectEx(filename, toa, kv, dateSub)
			report = 'OK! File đã được chỉnh sửa, các cells đã chỉnh sửa được bôi xanh. Hãy kiểm tra, chỉnh sửa bổ sung và nhấn "Ctrl+S" để lưu file "outCompare.xlsx".'
			return fO

		except PermissionError:
			report = "ERROR! Có thể file 'outCompare.xlsx' đang được mở, hãy đóng và thử lại."
		except FileNotFoundError:
			report = 'ERROR! Hãy chọn file cần chỉnh sửa bằng "Load Excel File...".'

		except NameError:
			report = 'ERROR! Hãy nhập tên toàn và khu vực.'
		
		self.statusBox.setText(report)

	def correctMdbFilesClick(self):
		report = "File đang được xử lí...."
		fI = self.pathMdbLine.text()

		fO = self.autoCorrectFileMdb()
		print(fO)
		try:
			compareEx(fI, fO)
			report = '''OK! File đã được chỉnh sửa!. Hãy kiểm tra lại.
						\n*Chú ý: 
						\n- File 'outCompare.xlsx' sẽ được tự động lưu và mở lên.
						\n- Các cells đã chỉnh sửa được bôi xanh.
						\n- Hãy kiểm tra, chỉnh sửa bổ sung và nhấn "Ctrl+S" lưu những thay đổi(nếu có).
						\n- Nhấn nút Save bên dưới để thực hiện lưu file.
						'''

		except PermissionError:
			report = "ERROR! Có thể file 'outCompare.xlsx' đang được mở, hãy đóng và thử lại."
		except FileNotFoundError:
			report = 'ERROR! Hãy chọn file cần chỉnh sửa bằng "Load Excel File...".'


		self.statusBox.setText(report)
		return fO



	# HÀM XỬ EDIT EXCEL

	def loadExClick(self):
		fileName, _ = QFileDialog.getOpenFileName(self, "Open the file",
		'', "Excel file (*.xlsx);;All Files (*)")

		if fileName == "":
			report = " Bạn chưa chọn file muốn chỉnh sửa, hãy nhấn 'Load Excel File....' để chọn."
		else:
			report = '''
						\nFile: '%s' được load thành công!
						\n- Hãy điền đầy đủ các thông tin bên trên.
						\n- Nhấn nút "AutoCorrect" để thực hiện việc chỉnh sửa tự động.
						''' % fileName.upper()
		self.statusBox.setText(report)

		self.pathExLine.setText(fileName)
		print(fileName)

	def saveExOut(self):
		try:
			self.info = (self.toaLine.text().upper(), self.kvLine.text().upper(),self.DesLine.text() ,self.dateLine.text()[-5:-3], self.dateLine.text()[-2:], self.dateLine.text())
			print(self.info[5])
			fileNameSave, _ = QFileDialog.getSaveFileName(self,"saveFlle","%s동%s_계단_%s(%s%s).xlsx" % self.info[0:5] ,filter ="File Excel (*.xlsx *.)")
			fileAC = "outCompare.xlsx"
			saveFileEx(fileAC,self.info, fileNameSave)
			report = 'OK! File "%s" được lưu thành công.' % fileNameSave
		except PermissionError:
			report = "ERROR! Có thể file %s đang mở, hãy đóng và thử lại lần nữa." % fileNameSave
		except TypeError:
			report = 'Bạn đã hủy việc lưu file. File lưu không thành công.'

		except FileNotFoundError:
			report = 'ERROR! Hãy thực hiện việc sửa file bằng cách nhấn "AutoCorrect" trước khi lưu file.'

		self.statusBox.setText(report)
		# print(fileNameSave)

	def autoCorrectFileEx(self):
		filename = self.pathExLine.text()
		toa = self.toaLine.text()
		kv = self.kvLine.text()
		dateSub = self.dateLine.text()
		
		if toa =='':
			print("Check")

		try:
			fO = autoCorrectEx(filename, toa, kv, dateSub)
			report = 'OK! File đã được chỉnh sửa, các cells đã chỉnh sửa được bôi xanh. Hãy kiểm tra, chỉnh sửa bổ sung và nhấn "Ctrl+S" để lưu file "outCompare.xlsx".'
			return fO

		except PermissionError:
			report = "ERROR! Có thể file 'outCompare.xlsx' đang được mở, hãy đóng và thử lại."
		except FileNotFoundError:
			report = 'ERROR! Hãy chọn file cần chỉnh sửa bằng "Load Excel File...".'

		except NameError:
			report = 'ERROR! Hãy nhập tên toàn và khu vực.'
		except TypeError:
			report = 'ERROR! Form Mẫu file excel cần sửa sai mẫu.'
		
		self.statusBox.setText(report)

	def correctExFilesClick(self):
		report = "File đang được xử lí...."
		fI = self.pathExLine.text()

		fO = self.autoCorrectFileEx()
		print(fO)
		try:
			compareEx(fI, fO)
			report = '''OK! File đã được chỉnh sửa!. Hãy kiểm tra lại.
						\n*Chú ý: 
						\n- File 'outCompare.xlsx' sẽ được tự động lưu và mở lên.
						\n- Các cells đã chỉnh sửa được bôi xanh.
						\n- Hãy kiểm tra, chỉnh sửa bổ sung và nhấn "Ctrl+S" lưu những thay đổi(nếu có).
						\n- Nhấn nút Save bên dưới để thực hiện lưu file.
						'''

		except PermissionError:
			report = "ERROR! Có thể file 'outCompare.xlsx' đang được mở, hãy đóng và thử lại."
		except FileNotFoundError:
			report = 'ERROR! Hãy chọn file cần chỉnh sửa bằng "Load Excel File...".'

		except:
			report = "ERROR! ERROR! Kiểm tra lại form mẫu file excel cần sửa, có thể file sai mẫu.!"

		self.statusBox.setText(report)
		return fO


	# HAM INFO:

	def chooseDes(self):
		self.DesLine.setText(listDesAV[self.Des2Combo.currentText()])



	## HÀM Help

	def clipClicked(self):
		print("clipButton clicked")
		webbrowser.open("https://i.pinimg.com/originals/d4/7d/6c/d47d6c686213e7579eb28f891f0d8ec2.jpg")
		pass	

	def docClicked(self):
		print("docButton clicked")
		webbrowser.open("https://www.facebook.com/photo.php?fbid=1377936218945555&set=rpd.100001875694283&type=3&theater")
		pass



	## HÀM ĐIỂU KHIỂN APP

	def runDataSet(self):
		try:
			os.startfile("C:/HD-FORM2015/DataSetting.exe")
			report = 'Data Setting is running.'
		except FileNotFoundError:
			report = 'Error! Chưa cài HD-FORM2015!.'
		self.statusBox.setText(report)

	def quitReClicked(self):
		for tempfile in ['autoCorrected Output File.xlsx','outCompare.xlsx']:
			try:
				os.remove(tempfile)
				report = "File %s tạm đã được xóa!" % tempfile
			except PermissionError:
				report = "ERROR! Có thể file 'outCompare.xlsx' đang được mở, hãy đóng và thử lại."
				break
				# self.statusBox.setText(report)
			except FileNotFoundError:
				report = 'Các file tạm đã được xóa!.'
		self.statusBox.setText(report)

	def quitClicked(self):
		self.close()

# Check Device:

class CheckDevice(QWidget):
	"""docstring for CheckDevice"""
	def __init__(self, arg):
		super(CheckDevice, self).__init__()
		self.arg = arg
		
def checkingDevice():
	res = requests.get("https://raw.githubusercontent.com/namtrung205/JNB-tools/master/theSupportList")
	theList = res.text.split(",")
	print(theList)
	regFile = open('C:/HDFormReg.dat', 'r')
	subList = []
	for line in regFile:
		subList.append(line.split('|'))
	# print(subList)
	for i in subList:
		if i[0] == 'AUTHCODE' and i[1][0:10] in theList:
			return True
		else:
			pass
	return False

# Autorun program 

if __name__ == '__main__':
	import sys
	from PyQt5.QtWidgets import QApplication
	app = QApplication(sys.argv)
	app.setWindowIcon(QIcon('Icon.ico'))

	try:
		if checkingDevice():
			JnbTool = EditExcelFile()
			JnbTool.show()
		else:
			message = QMessageBox()
			message.setWindowTitle("Error!!!")
			message.setText("Error! Your device is not in the list!")
			message.show()
			app.exit()
	except:
		message = QMessageBox()
		message.setWindowTitle("Error!!!")
		message.setText("Error! ConnectionError or unknowError!")
		message.show()
		app.exit()

	sys.exit(app.exec_())
